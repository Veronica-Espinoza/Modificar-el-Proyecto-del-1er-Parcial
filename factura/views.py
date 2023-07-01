from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from factura.models import Factura
from factura.forms import FacturaFormulario
from openpyxl import Workbook

# Create your views here.
def agregar_factura(request):
    pagina = loader.get_template('agregar_factura.html')
    if request.method == 'GET':
        formulario= FacturaFormulario
    elif request.method == 'POST':
        formulario= FacturaFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_factura(request, idFactura):
    pagina= loader.get_template('ver_factura.html')
    #persona= Persona.objects.get(pk=idPersona)
    factura= get_object_or_404(Factura, pk=idFactura)
    mensaje= {'factura': factura}
    return HttpResponse(pagina.render(mensaje, request))

def editar_factura(request, idFactura):
    pagina= loader.get_template('editar_factura.html')
    #persona= Persona.objects.get(pk=idPersona)
    factura= get_object_or_404(Factura, pk=idFactura)
    if request.method == 'GET':
        formulario= FacturaFormulario(instance=factura)
    elif request.method == 'POST':
        formulario= FacturaFormulario(request.POST, instance=factura)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    mensaje= {'formulario': formulario}
    return HttpResponse(pagina.render(mensaje, request))

def eliminar_factura(request, idFactura):
    factura= get_object_or_404(Factura, pk=idFactura)
    if factura:
        factura.delete()
        return redirect('inicio')

def generar_reporte (request):
    # Obtenemos todas las personas de nuestra base de datos
    factura = Factura.objects.all()
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE FACTURAS'
    ws['B1'] = 'REPORTE DE FACTURAS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'CODIGO_FACTURA'
    ws['C3'] = 'FECHA'
    ws['D3'] = 'CEDULA'
    ws['E3'] = 'NOMBRE'
    ws['F3'] = 'APELLIDO'
    ws['G3'] = 'EMAIL'
    ws['H3'] = 'PRODUCTO'
    ws['I3'] = 'CANTIDAD'
    ws['J3'] = 'PRECIO'
    ws['K3'] = 'TOTAL'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for factura in factura :
        ws.cell(row=cont, column=2).value = factura.id
        ws.cell(row=cont, column=3).value = factura.fecha
        ws.cell(row=cont, column=4).value = factura.cliente.cedula
        ws.cell(row=cont, column=5).value = factura.cliente.nombre
        ws.cell(row=cont, column=6).value = factura.cliente.apellido
        ws.cell(row=cont, column=7).value = factura.cliente.email
        ws.cell(row=cont, column=8).value = factura.producto.nombre
        ws.cell(row=cont, column=9).value = factura.cantidad
        ws.cell(row=cont, column=10).value = factura.producto.precio
        ws.cell(row=cont, column=11).value = factura.total
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteFacturasExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
